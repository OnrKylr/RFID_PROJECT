# -*- coding: utf-8 -*-

import serial
import openpyxl
import keyboard  # keyboard kütüphanesi
from datetime import datetime

ser = serial.Serial('COM7', 9600)


kartlar = {
    "17918779169": "vasfi bahtiyar",
    "147100236166": "Durhan Kamer",
    "9970209207": "Ayşe",
    "345147172": "Fatma",
    "11111": "Ali"
}

kart_durum = {k: False for k in kartlar.keys()}  # Kart durumlarını tutan sözlük
kart_baslama_zamani = {k: None for k in kartlar.keys()}  # Kart başlama zamanlarını tutan sözlük


def veriyi_kaydet(rfid_number, isim, durum):
    excel_dosyasi = "data.xlsx"
    wb = openpyxl.load_workbook(excel_dosyasi)
    ws = wb.active

    son_satir = ws.max_row + 1
    tarih = datetime.now().strftime('%Y-%m-%d')
    saat = datetime.now().strftime('%H:%M:%S')

    ws.cell(row=son_satir, column=1, value=rfid_number)
    ws.cell(row=son_satir, column=2, value=isim)
    ws.cell(row=son_satir, column=3, value=durum)
    ws.cell(row=son_satir, column=4, value="MKK Numarası")  # MKK numarası eksik
    ws.cell(row=son_satir, column=5, value=tarih)
    ws.cell(row=son_satir, column=6, value=saat)
    ws.cell(row=son_satir, column=7, value="Çalışma Süresi")  # Çalışma süresi eksik

    # Çalışma süresini hesapla ve excele yaz
    if kart_baslama_zamani[rfid_number] is not None and durum == "Durdu":
        baslama_zamani = kart_baslama_zamani[rfid_number]
        durma_zamani = datetime.now()
        calisma_suresi = durma_zamani - baslama_zamani
        ws.cell(row=son_satir, column=7, value=str(calisma_suresi))

        kart_baslama_zamani[rfid_number] = None
    elif durum == "Başladı":
        kart_baslama_zamani[rfid_number] = datetime.now()

    wb.save(excel_dosyasi)

    # Ekrana yazdırma
    print(f"KART NUMARASI: {rfid_number}")
    print(f"İSİM: {isim}")
    print(f"DURUM: {durum}")
    print(f"MKK: MKK Numarası")  # MKK numarası eksik
    print(f"TARİH: {tarih}")
    print(f"SAAT: {saat}")
    if kart_baslama_zamani[rfid_number] is not None and durum == "Durdu":
        print(f"ÇALIŞMA SÜRESİ: {calisma_suresi}")
    print()


def restart_esp8266():
    ser.write("restart\n".encode())  # "restart" komutunu ESP8266'a gönder


def kayit_dosyasi_olustur(veri):
    with open("kayit.txt", "w") as dosya:
        dosya.write(veri)


def kayit_dosyasi_oku():
    try:
        with open("kayit.txt", "r") as dosya:
            return dosya.read()
    except FileNotFoundError:
        return ""


def main():
    kayitli_veri = kayit_dosyasi_oku()  # Kaydedilmiş veriyi oku
    if kayitli_veri:
        print("Kaydedilen veri:", kayitli_veri)

    while True:
        if ser.in_waiting > 0:
            try:
                line = ser.readline().strip().decode('utf-8')
                if line in kartlar:
                    rfid_number = line
                    isim = kartlar[line]
                    print("RFID Kart Numarası:", rfid_number)
                    print("Kişi:", isim)

                    kart_durumu = kart_durum[rfid_number]
                    if kart_durumu:
                        durum = "Durdu"
                    else:
                        durum = "Başladı"

                    kart_durum[rfid_number] = not kart_durumu
                    veriyi_kaydet(rfid_number, isim, durum)
                else:
                    print("Geçersiz kart numarası:", line)
            except UnicodeDecodeError:
                print("Hatalı veya beklenmeyen karakter")
                continue

        # Klavyeden herhangi bir tuşa basıldığında "restart" komutunu gönder
        if keyboard.is_pressed(" "):
            restart_esp8266()

        # Güncel veriyi kaydederek, her döngüde yeniden yazma
        kayit_dosyasi_olustur(kayitli_veri)


if __name__ == "__main__":
    main()